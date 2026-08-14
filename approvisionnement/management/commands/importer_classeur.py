from pathlib import Path

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from approvisionnement.models import Acteur, Fonctionnalite, Module


class Command(BaseCommand):
    help = 'Importe les fonctionnalités du fichier Classeur1.xlsx'

    def add_arguments(self, parser):
        parser.add_argument(
            'fichier',
            nargs='?',
            default=str(Path('Classeur1.xlsx')),
        )

    def handle(self, *args, **options):
        chemin = Path(options['fichier'])
        if not chemin.is_file():
            chemin = Path(__file__).resolve().parents[3] / 'Classeur1.xlsx'
        if not chemin.is_file():
            self.stderr.write(self.style.ERROR(f'Fichier introuvable : {chemin}'))
            return

        wb = load_workbook(chemin, data_only=True)
        feuille = wb.active
        lignes = list(feuille.iter_rows(min_row=2, values_only=True))

        crees, maj = 0, 0
        for row in lignes:
            code, module_nom, libelle, acteurs, comportement, donnees, regles, priorite = (
                row + (None,) * 8
            )[:8]
            if not code:
                continue

            module, _ = Module.objects.get_or_create(nom=str(module_nom).strip())
            fonctionnalite, created = Fonctionnalite.objects.update_or_create(
                code=str(code).strip(),
                defaults={
                    'module': module,
                    'libelle': str(libelle or '').strip(),
                    'comportement_attendu': str(comportement or '').strip(),
                    'donnees_principales': str(donnees or '').strip(),
                    'regles_gestion': str(regles or '').strip(),
                    'priorite': str(priorite or '').strip(),
                },
            )
            noms_acteurs = [
                nom.strip()
                for nom in str(acteurs or '').replace(',', '/').split('/')
                if nom.strip()
            ]
            ids = []
            for nom in noms_acteurs:
                acteur, _ = Acteur.objects.get_or_create(nom=nom)
                ids.append(acteur.pk)
            fonctionnalite.acteurs.set(ids)
            if created:
                crees += 1
            else:
                maj += 1

        self.stdout.write(
            self.style.SUCCESS(
                f'Import terminé : {crees} créée(s), {maj} mise(s) à jour.'
            )
        )
